# core/conversion/handlers/excel_handler.py
import pandas as pd
import os
from pathlib import Path
import openpyxl
from typing import Dict, Tuple, Optional, List
from core.conversion.handlers.common.main_code import process_df_with_suffix


def process_excel(input_file: str,
                  map_dict: Dict[Tuple[str, str, str], List[Tuple[str, str, str, str]]], 
                  address_groups=None,
                  pool=None,
                ) -> bool:
    """
    Xử lý file Excel (.xlsx, .xlsm, .xls) – **không chuẩn hóa dữ liệu**.
    Đọc toàn bộ file bằng openpyxl (read_only + data_only) → tạo DataFrame → mapping → lưu.
    """
    # -------------------------------------------------
    # 0. KIỂM TRA CÁC TRƯỜNG ĐỊA CHỈ
    # -------------------------------------------------
    if not address_groups:
        print("Không có nhóm nào được chọn")
        return False

    # -------------------------------------------------
    # 1. KIỂM TRA FILE
    # -------------------------------------------------
    p = Path(input_file)
    if not p.exists():
        print(f"File Excel không tồn tại: {input_file}")
        return False

    # -------------------------------------------------
    # 2. ĐỌC EXCEL
    # -------------------------------------------------
    wb = None
    try:
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        sheet = wb.active

        if sheet.max_row <= 1:
            print("File Excel rỗng hoặc chỉ có header")
            wb.close()
            return False

        # ---- Header ----
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        header = [
            str(cell) if cell is not None else f"col_{i}"
            for i, cell in enumerate(header_row)
        ]
        
        # ---- Dữ liệu  ----
        data_rows = []
        total_rows = sheet.max_row - 1
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            clean_row = ['' if cell is None else cell for cell in row]
            data_rows.append(clean_row)

    except Exception as e:
        print(f"Lỗi đọc file Excel bằng openpyxl: {e}")
        if wb:
            wb.close()
        return False
    finally:
        if wb:
            wb.close()

    if not data_rows:
        print("Không có dữ liệu để xử lý")
        return False

    df = pd.DataFrame(data_rows, columns=header)
    print(f"📊 Đã đọc Excel: {len(df)} mẫu, {len(df.columns)} trường")

    if 'Trạng thái chuyển đổi' not in df.columns:
        df.insert(len(df.columns), 'Trạng thái chuyển đổi', '')

    # -------------------------------------------------
    # 3. XỬ LÝ DATAFRAME
    # -------------------------------------------------
    for idx, (id_p, id_d, id_w, p, d, w) in enumerate(address_groups):
        suffix = f"_group{idx+1}"
        df = process_df_with_suffix(df, map_dict,
                                    id_province_col=id_p, 
                                    id_district_col=id_d, 
                                    id_ward_col=id_w,
                                    province_col=p, 
                                    district_col=d, 
                                    ward_col=w,
                                    suffix=suffix, 
                                    pool=pool)
        
    count_success = (df['Trạng thái chuyển đổi'] == 'Thành công').sum()
    count_fail = len(df) - count_success

    df.insert(0, 'id_VNA', df.index + 1)
    
    return {
        "success": True,
        "full_df": df.to_dict(orient="records"),
        "total_rows": len(df),
        "success_count": int(count_success),
        "fail_count": int(count_fail),
    }